import os
import re
from typing import Optional, Dict, List, Any, Tuple
from datetime import date
import sys
import pandas as pd
from enum import Enum
from dateutil.parser import parse
import openpyxl
from openpyxl import Workbook  
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import locale
import json
import warnings
import time

class DocumentColumns(Enum):
    REFERENCE_NUMBER = 0,
    BILL_DATE = 1,
    AMOUNT_DUE = 2,
    BILL_TYPE = 3

class BillType(Enum):
    Bill = 0,
    Credit = 1

    @classmethod
    def from_string(cls, bill_type_str: str):
        if bill_type_str == 'Bill' or bill_type_str == 'Invoice':
            return cls.Bill
        elif bill_type_str == 'Credit' or bill_type_str == 'Credit Memo':
            return cls.Credit
        else:
            return None

class Document:
    def __init__(self, reference_number: str, bill_date: date, amount_due: float, bill_type: BillType):
        self.reference_number = reference_number
        self.bill_date = bill_date
        self.amount_due = amount_due
        self.bill_type = bill_type
    
    def __str__(self) -> str:
        return f"Reference Number: {self.reference_number}, Bill Date: {self.bill_date}, Amount Due: {self.amount_due}, Bill Type: {self.bill_type}"
    
    def __eq__(self, __value: object) -> bool:
        if not isinstance(__value, Document):
            return False  # Not the same type, can't be equal 


        return (
                self.reference_number == __value.reference_number and 
                self.bill_date == __value.bill_date and 
                self.amount_due == __value.amount_due and
                self.bill_type == __value.bill_type
            )

def getPath(root_path: str, file_name_pattern: str) -> Optional[str]:
    for filename in os.listdir(root_path): 
        if re.search(re.escape(file_name_pattern), filename):
            file_path = os.path.join(root_path, filename)
            return file_path
        
    return None

def processDocuments(fs_file_path, ir_file_path, output_file_name):
    print("\n-------------->\nProcessing: ")
    print(f"Search File at {FS_path}")
    print(f"Statement File at {IR_path}")

    FS_data_frame = pd.read_excel(fs_file_path)
    IR_data_frame = pd.read_excel(ir_file_path)

    print("Reading files...")

    FS_data_dict: Dict[str, List[Any]] = FS_data_frame.to_dict()
    IR_data_dict: Dict[str, List[Any]] = IR_data_frame.to_dict()

    FS_key_for_comparison = set(['ReferenceNumber', 'BillDate', 'AmountDue', 'BillType'])
    IR_key_for_comparison = set(['Document Number', 'Document Date', 'Amount Due', 'Document Type'])

    FS_data_dict = {key: FS_data_dict[key] for key in FS_key_for_comparison}
    IR_data_dict = {key: IR_data_dict[key] for key in IR_key_for_comparison}

    document_columns_to_FS_keys_dict = dict([(DocumentColumns.REFERENCE_NUMBER, 'ReferenceNumber'), (DocumentColumns.BILL_DATE, 'BillDate'), (DocumentColumns.AMOUNT_DUE, 'AmountDue'), (DocumentColumns.BILL_TYPE, 'BillType')])
    document_columns_to_IR_keys_dict = dict([(DocumentColumns.REFERENCE_NUMBER, 'Document Number'), (DocumentColumns.BILL_DATE, 'Document Date'), (DocumentColumns.AMOUNT_DUE, 'Amount Due'), (DocumentColumns.BILL_TYPE, 'Document Type')])

    FS_bill_type_key_to_bill_type_dict = dict([('Bill', BillType.Bill), ('Credit', BillType.Credit)])
    IR_bill_type_key_to_bill_type_dict = dict([('Invoice', BillType.Bill), ('Credit Memo', BillType.Credit)])

    document_column_to_output_key_dict = dict([(DocumentColumns.REFERENCE_NUMBER, 'Reference Number'), (DocumentColumns.BILL_DATE, 'Bill Date'), (DocumentColumns.AMOUNT_DUE, 'Amount Due'), (DocumentColumns.BILL_TYPE, 'Bill Type')])
    document_bill_type_to_output_value_dict = dict([(BillType.Bill, 'Bill'), (BillType.Credit, 'Credit')])

    def parse_reference_number(reference_number) -> str:
        try:
            # Try converting to float directly
            float_value = float(reference_number)

            # Handle integers and floats with decimal parts
            if float_value.is_integer():
                return str(int(float_value))  # Remove trailing ".0" if integer 
            else:
                return str(float_value)

        except (ValueError, TypeError):
            # Handle non-numeric values 
            return str(reference_number)

    def parseData(data: Dict[str, List[Any]], document_columns_to_data_key_dict: Dict[DocumentColumns, str], bill_type_to_data_key: Dict[BillType, str]) -> List[Document]:
        reference_numbers = list(data[document_columns_to_data_key_dict[DocumentColumns.REFERENCE_NUMBER]].values())
        bill_dates = list(data[document_columns_to_data_key_dict[DocumentColumns.BILL_DATE]].values())
        amount_dues = list(data[document_columns_to_data_key_dict[DocumentColumns.AMOUNT_DUE]].values())
        bill_types = list(data[document_columns_to_data_key_dict[DocumentColumns.BILL_TYPE]].values())

        documents: List[Document] = []

        for i in range(len(reference_numbers)):
            reference_number = parse_reference_number(reference_numbers[i])
            if reference_number == 'nan':
                continue

            bill_date =  bill_dates[i].date() if isinstance(bill_dates[i], pd.Timestamp) else parse(bill_dates[i]).date()
            amount_due = amount_dues[i]
            bill_type = bill_type_to_data_key[bill_types[i]]

            documents.append(Document(reference_number, bill_date, amount_due, bill_type))

        return documents


    FS_documents: List[Document] = parseData(FS_data_dict, document_columns_to_FS_keys_dict, FS_bill_type_key_to_bill_type_dict)
    IR_documents: List[Document] = parseData(IR_data_dict, document_columns_to_IR_keys_dict, IR_bill_type_key_to_bill_type_dict)

    print("Converted files to Document objects...")

    FS_documents_dict: Dict[str, Document] = {document.reference_number: document for document in FS_documents}

    output_documents: List[Tuple[Document, Optional[Document]]] = []

    for IR_document in IR_documents:
        FS_document = FS_documents_dict.get(IR_document.reference_number)

        if FS_document is None:
            output_documents.append((None, IR_document))
            continue

        if IR_document != FS_document:
            output_documents.append((FS_document, IR_document))

    print("Found differences...")

    reference_key = document_column_to_output_key_dict[DocumentColumns.REFERENCE_NUMBER]
    bill_date_key = document_column_to_output_key_dict[DocumentColumns.BILL_DATE]
    amount_due_key = document_column_to_output_key_dict[DocumentColumns.AMOUNT_DUE]
    bill_type_key = document_column_to_output_key_dict[DocumentColumns.BILL_TYPE]

    def format_as_currency(number: float) -> float:
        return locale.format_string('%.2f', number, grouping=True)

    def transform_data(data: List[Tuple[Document, Optional[Document]]]) -> Tuple[List[Tuple[int, int]], List[Dict[str, Any]]]:
        rows = []
        indices = []

        row_index = 0
        for FS_item, IR_item in data:
            row = {
                f"Search Result {reference_key}": FS_item.reference_number if FS_item is not None else "Not Found",
                f"IR Statement {reference_key}": IR_item.reference_number if IR_item is not None else "Not Found",
                f"Search Result {bill_date_key}": FS_item.bill_date if FS_item is not None else "Not Found",
                f"IR Statement {bill_date_key}": IR_item.bill_date if IR_item is not None else "Not Found",
                f"Search Result {amount_due_key}": format_as_currency(FS_item.amount_due) if FS_item is not None else "Not Found",
                f"IR Statement {amount_due_key}": format_as_currency(IR_item.amount_due) if IR_item is not None else "Not Found",
                f"Search Result {bill_type_key}": document_bill_type_to_output_value_dict[FS_item.bill_type] if FS_item is not None else "Not Found",
                f"IR Statement {bill_type_key}": document_bill_type_to_output_value_dict[IR_item.bill_type] if IR_item is not None else "Not Found",
            }

            rows.append(row)

            if FS_item is None or FS_item.reference_number != IR_item.reference_number:
                indices.append((row_index, 0))
                indices.append((row_index, 1))

            if FS_item is None or FS_item.bill_date != IR_item.bill_date:
                indices.append((row_index, 2))
                indices.append((row_index, 3))

            if FS_item is None or FS_item.amount_due != IR_item.amount_due:
                indices.append((row_index, 4))
                indices.append((row_index, 5))
            
            if FS_item is None or FS_item.bill_type != IR_item.bill_type:
                indices.append((row_index, 6))
                indices.append((row_index, 7))

            row_index += 1

        return (indices, rows)

    indices_with_difference, transformed_data = transform_data(output_documents)

    def convert_coordinates_to_excel_refs(coordinates):
        """Converts a list of (row, col) tuples (0-indexed) to Excel cell references.

        Args:
            coordinates (List[Tuple[int, int]]): A list of (row, col) coordinates.

        Returns:
            List[str]: A list of corresponding Excel cell references.
        """

        excel_refs = []
        for row, col in coordinates:
            col_letter = openpyxl.utils.get_column_letter(col + 1)  # Adjust for 1-based indexing
            excel_ref = col_letter + str(row + 2)  # Adjust for 1-based indexing + header
            excel_refs.append(excel_ref)

        return excel_refs

    cell_references = convert_coordinates_to_excel_refs(indices_with_difference)

    output_data_frame = pd.DataFrame(transformed_data)

    output_data_frame.to_excel(f"{output_file_name}.xlsx", index=False)  
    
    print("Created excel document with differences...")
    
    def highlight_cells_in_excel(worksheet, cell_references, color_rgb: str):
        for cell in cell_references:
            worksheet[cell].font = Font(color=color_rgb)

    def auto_with_adjusting_in_excel(worksheet):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column number
            for cell in col:
                try:
                    # Find the maximum length of content in the column
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            
            # Adjust the column width based on the maximum content length
            adjusted_width = (max_length + 2) * 1.2  # Adjust the width with a factor
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    def center_alignment_in_excel(worksheet):
        for col in worksheet.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    workbook = openpyxl.load_workbook(f"{output_file_name}.xlsx")
    worksheet = workbook['Sheet1']
    
    highlight_cells_in_excel(worksheet, cell_references, "FF0000")
    auto_with_adjusting_in_excel(worksheet)
    center_alignment_in_excel(worksheet)

    print("Formated excel document with differences...")

    workbook.save(f"{output_file_name}.xlsx")

    print(f"Saved formated excel document as {output_file_name}")

start_time = time.time()

if getattr(sys, 'frozen', False):
    project_root = os.path.dirname(sys.executable)
else:
    project_root = os.path.dirname(__file__)

config_file_path = getPath(project_root, "config.json")

if config_file_path is None:
    print("Config file named 'config.json' has not been found.")
    input("Enter something to close...")
    sys.exit()

with open(config_file_path, 'r') as f:
    file_contents = f.read()
    try:
        data = json.loads(file_contents)
    except Exception as e :
        print("Check the correct format of the config file.")
        input("Enter something to close...")
        sys.exit()
    

for config in data:
    statement_file_pattern = config["statementFileName"]
    search_file_pattern = config["searchFileName"]
    output_file_name = config["outputFileName"]

    FS_path = getPath(project_root, search_file_pattern)
    IR_path = getPath(project_root, statement_file_pattern)

    if FS_path is None or IR_path is None:
        print(f"One or more files have not been found. Search File: {search_file_pattern}, Statement File: {statement_file_pattern}")
        input("Enter something to close...")
        sys.exit()

    warnings.filterwarnings("ignore", category=UserWarning, message="Cannot parse header or footer so it will be ignored")

    processDocuments(FS_path, IR_path, output_file_name)

end_time = time.time()
runtime = round(end_time - start_time, 2)

print(f"\nCongratulation! Your work has been done in {runtime} seconds :)")
input("Enter something to close...")